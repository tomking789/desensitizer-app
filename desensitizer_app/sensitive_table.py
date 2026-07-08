from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SENSITIVE_HEADERS = ("原文", "替换为", "启用", "备注")


@dataclass
class SensitiveTableRow:
    row_number: int
    value: str
    replacement: str
    enabled: bool = True
    note: str = ""
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class SensitiveExportRow:
    value: str
    replacement: str
    enabled: bool
    note: str = ""


def write_sensitive_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "敏感词模板"
    sheet.append(SENSITIVE_HEADERS)
    _style_header(sheet, len(SENSITIVE_HEADERS))
    widths = [30, 30, 12, 42]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    guide = workbook.create_sheet("填写说明")
    guide.append(["字段", "说明"])
    guide.append(["原文", "需要被替换的敏感词或原始文本，不能为空。"])
    guide.append(["替换为", "脱敏后的显示文本，例如 客户001、人员001、项目001，不能为空。"])
    guide.append(["启用", "可填写 是/否、1/0、true/false；留空默认为是。"])
    guide.append(["备注", "可选，仅用于自己说明，不参与脱敏。"])
    _style_header(guide, 2)
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 76
    workbook.save(path)


def write_sensitive_export(path: Path, rows: Iterable[SensitiveExportRow]) -> None:
    if path.suffix.lower() == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(SENSITIVE_HEADERS)
            for row in rows:
                writer.writerow([row.value, row.replacement, "是" if row.enabled else "否", row.note])
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "敏感词设置"
    sheet.append(SENSITIVE_HEADERS)
    _style_header(sheet, len(SENSITIVE_HEADERS))
    for row in rows:
        sheet.append([row.value, row.replacement, "是" if row.enabled else "否", row.note])
    widths = [32, 32, 12, 46]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(path)


def read_sensitive_table(path: Path) -> list[SensitiveTableRow]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError("仅支持导入 .xlsx 或 .csv 敏感词表。")


def _read_xlsx(path: Path) -> list[SensitiveTableRow]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    headers = [_normalize_header(cell) for cell in rows[0]]
    result: list[SensitiveTableRow] = []
    for index, row in enumerate(rows[1:], start=2):
        values = _row_to_dict(headers, row)
        if not any(str(item or "").strip() for item in row):
            continue
        result.append(_make_row(index, values))
    return result


def _read_csv(path: Path) -> list[SensitiveTableRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        reader = csv.reader(handle)
        try:
            first_row = next(reader)
        except StopIteration:
            return []
        headers = [_normalize_header(cell) for cell in first_row]
        if "value" not in headers:
            headers = ["value", "replacement", "enabled", "note"]
            data_rows = [first_row, *list(reader)]
            start = 1
        else:
            data_rows = list(reader)
            start = 2
        result: list[SensitiveTableRow] = []
        for offset, row in enumerate(data_rows, start=start):
            if not any(str(item or "").strip() for item in row):
                continue
            values = _row_to_dict(headers, row)
            result.append(_make_row(offset, values))
        return result


def _make_row(row_number: int, values: dict[str, object]) -> SensitiveTableRow:
    return SensitiveTableRow(
        row_number=row_number,
        value=str(values.get("value") or "").strip(),
        replacement=str(values.get("replacement") or "").strip(),
        enabled=_parse_enabled(values.get("enabled")),
        note=str(values.get("note") or "").strip(),
    )


def _row_to_dict(headers: list[str], row: Iterable[object]) -> dict[str, object]:
    values: dict[str, object] = {}
    for header, value in zip(headers, row):
        if header:
            values[header] = value
    return values


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    aliases = {
        "原文": "value",
        "敏感词": "value",
        "词汇": "value",
        "value": "value",
        "original": "value",
        "替换为": "replacement",
        "替换值": "replacement",
        "脱敏后": "replacement",
        "replacement": "replacement",
        "replace with": "replacement",
        "启用": "enabled",
        "是否启用": "enabled",
        "enabled": "enabled",
        "备注": "note",
        "说明": "note",
        "note": "note",
    }
    return aliases.get(text, text)


def _parse_enabled(value: object) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return True
    return text not in {"0", "false", "no", "n", "否", "停用", "禁用"}


def _style_header(sheet, column_count: int) -> None:
    fill = PatternFill("solid", fgColor="EAF1F7")
    for cell in sheet[1][:column_count]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = fill
